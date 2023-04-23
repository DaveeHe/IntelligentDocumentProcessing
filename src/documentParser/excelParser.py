# -*- coding: utf-8 -*-
# @Time : 2023/4/23 11:13
# @Author : Professor He
# @Email : hechangjia1018@163.com
# @File : excelParser.py
# @Project : IntelligentDocumentProcessing

import pandas as pd
import numpy as np
import openpyxl
from src.util.config_util import logger


def get_pandas_data(df):
    x, y = df.shape
    contents = []
    for dx in range(x):
        line = ''
        for dy in range(y):
            element = df.iloc[dx, dy]
            if not element:
                continue
            if element is np.NAN or str(element) == 'nan':
                continue
            line += str(element)
            line += '。'
        contents.append(line)
    return contents


def get_xlsx_data(path):
    workspaces = openpyxl.load_workbook(path)
    sheet_names = workspaces.get_sheet_names()
    contents = []
    for name in sheet_names:
        current_sheet = workspaces[name]
        rows, cols = current_sheet.max_row, current_sheet.max_column
        for row in range(1, rows+1):
            line = ''
            for col in range(1, cols+1):
                element = current_sheet.cell(row, col).value
                if element:
                    line += str(element)
                    line += '。'
            contents.append(line)

    return contents


def get_excel_data(path, filetype=''):
    contents = []
    if filetype == 'xlsx' or filetype == 'xls':
        try:
            data = pd.read_excel(path, header=None, sheet_name=None)
            for k, v in data.items():
                content = get_pandas_data(v)
                contents.extend(content)
        except Exception as e:
            logger(e.args)
            if filetype == 'xlsx':
                try:
                    contents = get_xlsx_data(path)
                except Exception as e:
                    logger(e.args)
    elif filetype == 'csv':
        try:
            data = pd.read_csv(path, header=None, encoding='utf-8')
            content = get_pandas_data(data)
            contents.extend(content)
        except Exception as e:
            logger(e.args)
            try:
                data = pd.read_csv(path, header=None, encoding='gb18030', error_bad_lines=False)
                content = get_pandas_data(data)
                contents.extend(content)
            except Exception as e:
                logger.info(e)
    return contents, []


if __name__ == '__main__':
    path = '/home/hechangjia/nlp_project/aiDocProcess/pra_data/1.xlsx'
    get_excel_data(path, filetype='xlsx')